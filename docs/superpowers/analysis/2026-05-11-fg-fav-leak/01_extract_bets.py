#!/usr/bin/env python3
"""Extract 627 MLB correlated parlays + join to placed_parlays edge_pct."""
import openpyxl, duckdb, re, sys
import pandas as pd
import numpy as np
from pathlib import Path

XLSX = Path('/tmp/bet_tracking.xlsx')
DASH_DB = Path('/Users/callancapitolo/NFLWork/Answer Keys/MLB Dashboard/mlb_dashboard.duckdb')
OUT = Path('/tmp/fg_fav_leak/bets.parquet')

# --- Load xlsx ---
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb['Sheet1']
rows = []
for r in range(2, ws.max_row + 1):
    sport = ws.cell(row=r, column=3).value
    bet_type = ws.cell(row=r, column=5).value
    desc = ws.cell(row=r, column=4).value
    if sport != 'MLB' or bet_type != 'Parlay' or not desc or 'PARLAY (2 TEAMS)' not in str(desc):
        continue
    rows.append({
        'date': ws.cell(row=r, column=1).value,
        'platform': ws.cell(row=r, column=2).value,
        'desc': str(desc),
        'odds_amer': ws.cell(row=r, column=7).value,
        'stake': ws.cell(row=r, column=8).value,
        'dec_odds': ws.cell(row=r, column=9).value,
        'result': ws.cell(row=r, column=10).value,
        'pnl': ws.cell(row=r, column=13).value,
    })
df = pd.DataFrame(rows)
df['date'] = pd.to_datetime(df['date'])
df['stake'] = pd.to_numeric(df['stake'], errors='coerce').fillna(0)
df['pnl'] = pd.to_numeric(df['pnl'], errors='coerce').fillna(0)
df['dec_odds'] = pd.to_numeric(df['dec_odds'], errors='coerce')

# --- Parse description ---
df['is_f5'] = df['desc'].apply(lambda d: bool(re.search(r'\b(1H|F5|1ST 5|FIRST 5)\b', d, re.IGNORECASE)))
df['slice'] = df['is_f5'].map({True: 'F5', False: 'FG'})

def line_type(d):
    n = d.replace('½', '.5').replace('1/2', '.5')
    if re.search(r'\-1\.5', n): return 'fav-1.5'
    if re.search(r'\+1\.5', n): return 'dog+1.5'
    if re.search(r'[+\-]\.?5', n): return 'half(F5)'
    return 'other'
df['line_type'] = df['desc'].apply(line_type)
df['ou'] = df['desc'].apply(lambda d: 'over' if re.search(r'TOTAL O', d, re.IGNORECASE) else ('under' if re.search(r'TOTAL U', d, re.IGNORECASE) else 'unk'))

def total_num(d):
    m = re.search(r'TOTAL [ou](\d+)([½.]?\d*)', d, re.IGNORECASE)
    if m:
        return int(m.group(1)) + (0.5 if ('½' in m.group(2) or '.5' in m.group(2)) else 0)
    return None
df['total_line'] = df['desc'].apply(total_num)

# Extract spread leg American odds. Pattern: `TEAM ±1½±NNN` or similar
def parse_spread_amer(d):
    n = d.replace('½', '.5')
    m = re.search(r'[+\-]1\.5\s*([+\-]\d+|EV)', n)
    if m:
        v = m.group(1).replace('EV', '+100')
        try: return int(v)
        except: return None
    return None
df['spread_amer'] = df['desc'].apply(parse_spread_amer)

# Extract total leg American odds. Pattern: `TOTAL [ou]N±NNN`
def parse_total_amer(d):
    n = d.replace('½', '.5')
    m = re.search(r'TOTAL [ou]\d+\.?5?\s*([+\-]\d+|EV)', n, re.IGNORECASE)
    if m:
        v = m.group(1).replace('EV', '+100')
        try: return int(v)
        except: return None
    return None
df['total_amer'] = df['desc'].apply(parse_total_amer)

# --- Teams ---
def teams(d):
    m = re.search(r'\(([^()]+?)\s+vrs\s+([^()]+?)\)', d, re.IGNORECASE)
    return (m.group(1).strip().upper(), m.group(2).strip().upper()) if m else ('', '')
df[['away_full', 'home_full']] = df['desc'].apply(lambda d: pd.Series(teams(d)))
df['week'] = df['date'].dt.to_period('W-SUN').apply(lambda p: p.start_time.date())

# --- Join to placed_parlays for edge_pct (only ~125 rows have this) ---
con = duckdb.connect(str(DASH_DB), read_only=True)
pp = con.execute("""
  SELECT parlay_hash, home_team, away_team, combo,
         spread_line, total_line as pp_total_line, fair_odds, wz_odds,
         edge_pct, actual_size, placed_at, ticket_number
  FROM placed_parlays
  WHERE status='placed' AND edge_pct IS NOT NULL
""").fetchdf()
con.close()

# Team-short aliases: xlsx uses different abbreviations for some clubs than dashboard.
# Map both sides to canonical short to avoid spurious mismatches.
SHORT_ALIAS = {
    'DIAMONDBACKS': 'DBACKS',  # xlsx writes "DBACKS"; dashboard writes "Arizona Diamondbacks"
}
def short(s):
    if pd.isna(s): return ''
    base = str(s).upper().split()[-1]
    return SHORT_ALIAS.get(base, base)
pp['home_short'] = pp['home_team'].apply(short)
pp['away_short'] = pp['away_team'].apply(short)
df['home_short'] = df['home_full'].apply(lambda s: short(s) if s else '')
df['away_short'] = df['away_full'].apply(lambda s: short(s) if s else '')
pp['date_key'] = pd.to_datetime(pp['placed_at']).dt.date

# Match: home_short + away_short + slice + ou + line_type + date ±2 days
def combo_to_meta(combo):
    is_f5 = combo.startswith('F5 ')
    ou = 'over' if 'Over' in combo else 'under'
    return is_f5, ou
pp[['is_f5', 'ou']] = pp['combo'].apply(lambda c: pd.Series(combo_to_meta(c)))
pp['slice'] = pp['is_f5'].map({True: 'F5', False: 'FG'})
def pp_lt(s):
    if abs(s) < 1: return 'half(F5)'
    return 'fav-1.5' if s < 0 else 'dog+1.5'
pp['line_type'] = pp['spread_line'].apply(pp_lt)

# Build matches. Sort pp by placed_at so earliest placements claim candidates first,
# and prefer an unused (still-NaN edge_pct) candidate before falling back to the first
# match — avoids losing rows when two placed parlays match the same bet on the same date.
df['edge_pct'] = np.nan
df['fair_odds'] = np.nan
df['wz_odds_pp'] = np.nan
df['parlay_hash'] = pd.Series([None] * len(df), dtype='object')
used_idx = set()
pp_sorted = pp.sort_values('placed_at')
for _, p in pp_sorted.iterrows():
    cand = df[(df['home_short'] == p['home_short']) & (df['away_short'] == p['away_short']) &
              (df['slice'] == p['slice']) & (df['ou'] == p['ou']) & (df['line_type'] == p['line_type'])]
    if len(cand) == 0: continue
    dd = (cand['date'].dt.date - p['date_key']).apply(lambda d: abs(d.days))
    cand = cand[dd <= 2]
    if len(cand) == 0: continue
    unused = [j for j in cand.index if j not in used_idx]
    j = unused[0] if unused else cand.index[0]
    used_idx.add(j)
    df.at[j, 'edge_pct'] = p['edge_pct']
    df.at[j, 'fair_odds'] = p['fair_odds']
    df.at[j, 'wz_odds_pp'] = p['wz_odds']
    df.at[j, 'parlay_hash'] = p['parlay_hash']

print(f"Loaded {len(df)} bets. Matched to placed_parlays: {df['edge_pct'].notna().sum()}")
print(f"Slice counts: {df['slice'].value_counts().to_dict()}")
print(f"FG-fav-1.5: n={((df['slice']=='FG')&(df['line_type']=='fav-1.5')).sum()}")
print(f"FG-fav-1.5 + over: n={((df['slice']=='FG')&(df['line_type']=='fav-1.5')&(df['ou']=='over')).sum()}")
print(f"FG-fav-1.5 + under: n={((df['slice']=='FG')&(df['line_type']=='fav-1.5')&(df['ou']=='under')).sum()}")

df.to_parquet(OUT)
print(f"Saved {OUT}")
